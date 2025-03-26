import requests
import pandas as pd
import time
import openpyxl
import geoip2.database
 
data = {}
head = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
 
def get_ip_region_from_online(ip):
    try:
        response = requests.get('http://ip.zxinc.org/api.php?type=json&ip='+ip , headers = head)
        text = response.json()
        
        location = text["data"]["location"]
 
        print(location)
        return(location)
    except Exception as e:
        print(e)
 
def get_ip_region(ip):
    """使用本地MaxMind数据库查询"""
    try:
        with geoip2.database.Reader("path-to-your-db\\GeoLite2-City.mmdb") as reader:
            resp = reader.city(ip)
            print(resp)
            return f"{resp.country.name} - {resp.subdivisions.most_specific.name} - {resp.city.name}"
    except Exception as e:
        return f"错误: {str(e)}"
 
def process_excel(filename):
    # 加载Excel文件
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # 定位列索引
    headers = [cell.value for cell in ws[1]]
    ip_col = headers.index("ip") + 1
    region_col = headers.index("IP-region") + 1
    
    # 逐行处理
    for row in ws.iter_rows(min_row=2):  # 从数据行开始
        ip_cell = row[ip_col-1]
        region_cell = row[region_col-1]
        
        if ip_cell.value and not region_cell.value:  # 需要处理的行
            region = get_ip_region(ip_cell.value)
            region_cell.value = region
            print(f"已处理 {ip_cell.value} -> {region}")
 
    wb.save(filename)  # 实时保存
    wb.close()
 
if __name__ == "__main__":
    process_excel("path-to-your-file")  # 替换为你的文件名
    print("处理完成！")
